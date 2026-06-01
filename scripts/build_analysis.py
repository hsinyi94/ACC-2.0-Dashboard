"""產出 ACC 2.0 分析表。

區塊 1:2.0 賣家單獨分析,資料來源 = 「2.0分析」工作表(既有樞紐表結果)。
區塊 2:與 1.0 賣家對比,從 ACC 1.0「最終錄取名單」原始資料重新計算,
        只對比 1.0 和 2.0 都有的項目。

設計決定:
- US vs Expansion 做為「錄取國家」附加列。
- Category 在 1.0 無對應,不做對比。
- 註冊商標題目版本不同,跳過對比。
- 1.0 站點欄若同一格多值(日本,中東)拆成各 +1,分母用賣家數。
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\hsinyih\OneDrive - amazon.com\ACC 2.0")
FILE_20 = BASE / "ACC 2.0 最終錄取結果 (20260105).xlsx"
FILE_10 = BASE / "ACC 1.0分析.xlsx"

OUT_DIR = Path(__file__).resolve().parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_FILE = OUT_DIR / f"ACC_2.0_分析_{dt.date.today().strftime('%Y%m%d')}.xlsx"


# ============================================================
# Step 1: 從「2.0分析」抓既有樞紐表區塊
# ============================================================

def load_20_analysis() -> pd.DataFrame:
    """讀 ACC 2.0 最終賣家名單 (113) 作為所有 2.0 分析的唯一來源。"""
    return pd.read_excel(FILE_20, sheet_name="最終賣家名單 (113)", engine="openpyxl")


# 新工作表欄位 -> 分析項目名稱的對照
BLOCKS_20_COLS: list[tuple[str, str]] = [
    ("公司類型", "公司類型"),
    ("公司所在區域", "公司所在區域"),
    ("Category", "Category"),
    ("公司產品類型", "公司產品類型"),
    ("公司主要經營型態", "公司主要經營型態"),
    ("是否有國內電商銷售經驗", "是否有國內「電商」銷售經驗"),
    ("品牌創立年限", "品牌創立年限"),
]


def get_20_blocks() -> dict[str, list[tuple[str, int, float]]]:
    """從最終賣家名單 (113) 直接計算各項目的分布。"""
    df = load_20_analysis()
    total = len(df)
    result: dict[str, list[tuple[str, int, float]]] = {}
    for item_name, col_name in BLOCKS_20_COLS:
        if col_name not in df.columns:
            print(f"  [警告] 找不到欄位 {col_name!r},跳過 {item_name}")
            continue
        vc = df[col_name].dropna().astype(str).str.strip().value_counts()
        rows = [(str(k), int(v), v / total) for k, v in vc.items()]
        result[item_name] = rows
    return result


def extract_block(
    df: pd.DataFrame,
    title_row: int,
    title_col: int,
    data_col_start: int,
    data_col_end: int,
) -> tuple[str, list[tuple[str, int, float]]]:
    """從 2.0分析 擷取一個樞紐區塊。

    title_row: 項目標題所在列 (例如「公司類型」)
    title_col: 項目標題所在欄
    data_col_start..data_col_end: 該區塊涵蓋的欄範圍
    回傳 (項目名稱, [(類別, 人數, 佔比), ...]),不含 Grand Total。
    """
    title = str(df.iat[title_row, title_col]).strip()
    # 標題下一列是 "Row Labels / Count / %"
    header_row = title_row + 1
    # 資料從 header_row + 1 開始,到遇到 Grand Total 為止
    rows: list[tuple[str, int, float]] = []
    for r in range(header_row + 1, len(df)):
        label = df.iat[r, data_col_start]
        if pd.isna(label):
            break
        label_s = str(label).strip()
        if label_s.lower().startswith("grand total"):
            break
        count_val = df.iat[r, data_col_start + 1]
        pct_val = df.iat[r, data_col_start + 2] if data_col_start + 2 <= data_col_end else None
        count = int(count_val) if pd.notna(count_val) else 0
        pct = float(pct_val) if pd.notna(pct_val) else 0.0
        rows.append((label_s, count, pct))
    return title, rows


# 根據 inspect 結果硬編定位 (見 dump_20_analysis.py 輸出)
# 格式: (item_name_override, title_row, title_col, data_col_start)
# 資料都是三欄寬 (Label, Count, %)
BLOCKS_20: list[tuple[str, int, int, int]] = [
    # 基本資訊
    ("公司類型", 4, 5, 5),
    ("公司所在區域", 15, 5, 5),
    ("Category", 25, 5, 5),
    ("公司產品類型", 34, 5, 5),
    ("公司主要經營型態", 55, 5, 5),
    # 基本條件
    ("是否有亞馬遜賣家帳號", 4, 9, 9),
    ("預計開始銷售的時間", 17, 9, 9),
    ("是否有國內電商銷售經驗", 15, 13, 13),
    # 加分條件
    ("品牌創立年限", 4, 13, 13),
    ("是否安排專人營運亞馬遜", 25, 13, 13),
    ("預計投入亞馬遜站內廣告預算", 4, 17, 17),
    ("預計投入亞馬遜FBA起始庫存", 15, 17, 17),
]


# ============================================================
# Step 2: 錄取國家 (AE 細分) 與 US vs Expansion
# ============================================================

def get_recruit_country() -> tuple[list[tuple[str, str, int, float]], list[tuple[str, int, float]]]:
    """從最終賣家名單 (113) 計算錄取國家分布。

    合併規則:AE+SA→MENA,UK+DE+FR→EU
    排序:US / EU / JP / MENA / AU

    回傳:
      ae_rows: [(國家, BD, 人數, 佔比), ...] 含小計列 (BD 欄為空)
      us_exp_rows: [('US', n, pct), ('Expansion (non-US)', n, pct)]
    """
    df = load_20_analysis()
    total = len(df)

    # 國家合併 mapping
    COUNTRY_MERGE = {
        "AE": "MENA",
        "SA": "MENA",
        "UK": "EU",
        "DE": "EU",
        "FR": "EU",
    }
    COUNTRY_ORDER = ["US", "EU", "JP", "MENA", "AU"]

    # 先做合併
    df["錄取國家_merged"] = df["錄取國家"].astype(str).str.strip().map(
        lambda x: COUNTRY_MERGE.get(x, x)
    )

    # 按合併後國家 + AE 分組
    ae_rows: list[tuple[str, str, int, float]] = []

    # 按指定順序
    country_counts = df["錄取國家_merged"].value_counts()
    for country in COUNTRY_ORDER:
        if country not in country_counts.index:
            continue
        country_total = int(country_counts[country])
        sub = df[df["錄取國家_merged"] == country]
        ae_vc = sub["AE"].dropna().astype(str).str.strip().value_counts()
        first = True
        for ae_name, ae_count in ae_vc.items():
            ae_rows.append((country if first else "", str(ae_name), int(ae_count), ae_count / total))
            first = False
        ae_rows.append((f"{country} 小計", "", country_total, country_total / total))

    # 其他未列入的國家
    for country in country_counts.index:
        if country in COUNTRY_ORDER or country == "nan":
            continue
        country_total = int(country_counts[country])
        sub = df[df["錄取國家_merged"] == country]
        ae_vc = sub["AE"].dropna().astype(str).str.strip().value_counts()
        first = True
        for ae_name, ae_count in ae_vc.items():
            ae_rows.append((country if first else "", str(ae_name), int(ae_count), ae_count / total))
            first = False
        ae_rows.append((f"{country} 小計", "", country_total, country_total / total))

    # US vs Expansion
    us_count = int(country_counts.get("US", 0))
    expansion_count = total - us_count
    us_exp_rows = [
        ("US", us_count, us_count / total if total else 0),
        ("Expansion (non-US)", expansion_count, expansion_count / total if total else 0),
    ]
    return ae_rows, us_exp_rows


# ============================================================
# Step 3: 1.0 對比計算
# ============================================================

def load_10_roster() -> pd.DataFrame:
    return pd.read_excel(FILE_10, sheet_name="最終錄取名單", engine="openpyxl")


# 站點中文 -> 國家代碼 (1.0 用區域級表達,括號裡是該區域明細,不拆)
REGION_MAP = {
    "北美": "US",  # 1.0「北美 (美國/加拿大/墨西哥)」對應 2.0 的 US 區
    "歐洲": "EU",
    "中東": "MENA",
    "日本": "JP",
    "澳洲": "AU",
    "新加坡": "SG",  # 2.0 無對應,保留以便檢查
}

_PAREN = re.compile(r"[（(][^）)]*[）)]")


def _parse_10_site(v: str) -> list[str]:
    """把一格站點值解析成區域代碼清單。例:
      '北美 (美國/加拿大/墨西哥),中東' -> ['US', 'MENA']
    """
    # 先拿掉所有括號內容 (含中英文括號)
    s = _PAREN.sub("", str(v))
    # 以中英文逗號拆
    parts = re.split(r"[,，]", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if p in REGION_MAP:
            out.append(REGION_MAP[p])
        else:
            out.append(p)  # 保留原值方便檢查
    return out


def count_10_country(df: pd.DataFrame) -> list[tuple[str, int, float]]:
    """1.0 的站點欄拆分計數,一格多值各 +1。"""
    counts: dict[str, int] = {}
    unknown: list[str] = []
    total_sellers = len(df)
    for v in df["站點"].dropna():
        for code in _parse_10_site(v):
            counts[code] = counts.get(code, 0) + 1
            if code not in {"US", "MENA", "JP", "EU", "AU", "SG"}:
                unknown.append(code)
    if unknown:
        print(f"  [警告] 1.0 站點有未對應的值: {set(unknown)}")
    order = ["US", "EU", "JP", "MENA", "AU", "SG"]
    rows = []
    for k in order:
        if k in counts:
            rows.append((k, counts[k], counts[k] / total_sellers))
    for k in sorted(counts):
        if k not in order:
            rows.append((k, counts[k], counts[k] / total_sellers))
    return rows


def count_10_us_vs_expansion(df: pd.DataFrame) -> list[tuple[str, int, float]]:
    """1.0 US vs Expansion:每位賣家若站點含 US 算 US,否則算 Expansion。"""
    total = len(df)
    us = 0
    exp = 0
    for v in df["站點"].dropna():
        codes = set(_parse_10_site(v))
        if "US" in codes:
            us += 1
        else:
            exp += 1
    return [
        ("US", us, us / total if total else 0),
        ("Expansion (non-US)", exp, exp / total if total else 0),
    ]


def count_column(df: pd.DataFrame, col: str) -> list[tuple[str, int, float]]:
    """對指定欄位做 value_counts,回傳 (類別, 人數, 佔比)。"""
    total = len(df)
    vc = df[col].dropna().astype(str).str.strip().value_counts()
    return [(str(k), int(v), v / total) for k, v in vc.items()]


# 1.0 可對比的欄位
# 使用 callable 而非欄位名稱,可自訂轉換邏輯 (如 1.0 品牌年限合併)
from typing import Callable

COMPARISON_MAP: list[tuple[str, str, Callable[[pd.DataFrame, str], list[tuple[str, int, float]]] | None]] = [
    # (2.0 項目, 1.0 欄位名, 自訂計算函式 or None=直接 value_counts)
    ("公司類型", "2-2. 公司類型", None),
    ("公司產品類型", "2-9.產品類型", None),
    ("公司主要經營型態", "2-5.公司主要經營型態", None),
    ("是否有國內電商銷售經驗", "2-11.是否有國內「電商」銷售經驗", None),
    ("品牌創立年限", "2-8.品牌創立年限", "brand_age"),  # 特殊處理:合併「六到十年+十年以上」
]


def count_10_brand_age(df: pd.DataFrame, col: str) -> list[tuple[str, int, float]]:
    """1.0 品牌年限:把「六到十年 + 十年以上」合併為「六年以上 (1.0 合併)」。

    這樣可對應 2.0 的「六到八年 + 八年以上」。
    """
    total = len(df)
    vc = df[col].dropna().astype(str).str.strip().value_counts().to_dict()
    merged_count = vc.pop("六到十年", 0) + vc.pop("十年以上", 0)
    rows: list[tuple[str, int, float]] = []
    # 先放有對應的類別
    for k in ["五年以下"]:
        if k in vc:
            rows.append((k, vc.pop(k), vc.get(k, 0) / total if False else vc.get(k, 0)))
    # 重算(剛剛 pop 過了,用原始資料重查)
    raw = df[col].dropna().astype(str).str.strip().value_counts()
    rows = []
    if "五年以下" in raw.index:
        c = int(raw["五年以下"])
        rows.append(("五年以下", c, c / total))
    c_merged = int(raw.get("六到十年", 0)) + int(raw.get("十年以上", 0))
    if c_merged:
        rows.append(("六年以上 (1.0 合併)", c_merged, c_merged / total))
    # 其餘原始類別 (如「不知道」)
    for k, v in raw.items():
        if k in ("五年以下", "六到十年", "十年以上"):
            continue
        rows.append((str(k), int(v), int(v) / total))
    return rows


# ============================================================
# Step 4: 輸出 Excel
# ============================================================

HEADER_FILL = PatternFill("solid", fgColor="305496")
SUBHEADER_FILL = PatternFill("solid", fgColor="8EA9DB")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBTOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
EXPANSION_FILL = PatternFill("solid", fgColor="E2EFDA")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
THICK_BORDER = Border(*[Side(style="medium", color="305496")] * 4)


def _fmt_cell(ws, cell_ref: str, *, bold=False, fill=None, align="left", number_format=None, color=None):
    c = ws[cell_ref]
    c.font = Font(bold=bold, color=color or "000000", size=11, name="Calibri")
    if fill:
        c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if number_format:
        c.number_format = number_format
    c.border = BORDER


def write_section1(ws, start_row: int, blocks: dict, ae_rows, us_exp_rows) -> int:
    """寫區塊 1,回傳下一個可用列號。"""
    r = start_row
    # 大標
    ws.cell(row=r, column=1, value="區塊 1:ACC 2.0 錄取賣家分析 (n=115)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _fmt_cell(ws, f"A{r}", bold=True, fill=HEADER_FILL, color="FFFFFF", align="center")
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = THICK_BORDER
    r += 2

    # --- 錄取國家 ---
    r = _write_country_table(ws, r, ae_rows, us_exp_rows)
    r += 1

    # --- 其他項目 ---
    order = [
        "公司類型",
        "公司所在區域",
        "Category",
        "公司產品類型",
        "公司主要經營型態",
        "是否有亞馬遜賣家帳號",
        "預計開始銷售的時間",
        "是否有國內電商銷售經驗",
        "品牌創立年限",
        "是否安排專人營運亞馬遜",
        "預計投入亞馬遜站內廣告預算",
        "預計投入亞馬遜FBA起始庫存",
    ]
    for name in order:
        rows = blocks.get(name, [])
        r = _write_item_table(ws, r, name, rows)
        r += 1
    return r


def _write_country_table(ws, r, ae_rows, us_exp_rows):
    # 項目標題
    ws.cell(row=r, column=1, value="錄取國家 (依 AE)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _fmt_cell(ws, f"A{r}", bold=True, fill=SECTION_FILL, align="left")
    r += 1
    # header
    headers = ["國家", "AE", "人數", "佔比"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
        _fmt_cell(ws, f"{get_column_letter(i)}{r}", bold=True, fill=SUBHEADER_FILL, color="FFFFFF", align="center")
    r += 1
    for country, bd, cnt, pct in ae_rows:
        is_subtotal = bd == "" and country.endswith("小計")
        ws.cell(row=r, column=1, value=country)
        ws.cell(row=r, column=2, value=bd)
        ws.cell(row=r, column=3, value=cnt)
        ws.cell(row=r, column=4, value=pct)
        fill = SUBTOTAL_FILL if is_subtotal else None
        _fmt_cell(ws, f"A{r}", bold=is_subtotal, fill=fill)
        _fmt_cell(ws, f"B{r}", fill=fill)
        _fmt_cell(ws, f"C{r}", bold=is_subtotal, fill=fill, align="right")
        _fmt_cell(ws, f"D{r}", bold=is_subtotal, fill=fill, align="right", number_format="0.00%")
        r += 1
    # US vs Expansion 附加列
    ws.cell(row=r, column=1, value="附加:US vs Expansion")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _fmt_cell(ws, f"A{r}", bold=True, fill=EXPANSION_FILL, align="left")
    r += 1
    ws.cell(row=r, column=1, value="類別")
    ws.cell(row=r, column=2, value="")
    ws.cell(row=r, column=3, value="人數")
    ws.cell(row=r, column=4, value="佔比")
    for i in (1, 3, 4):
        _fmt_cell(ws, f"{get_column_letter(i)}{r}", bold=True, fill=SUBHEADER_FILL, color="FFFFFF", align="center")
    _fmt_cell(ws, f"B{r}", fill=SUBHEADER_FILL)
    r += 1
    for label, cnt, pct in us_exp_rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value="")
        ws.cell(row=r, column=3, value=cnt)
        ws.cell(row=r, column=4, value=pct)
        _fmt_cell(ws, f"A{r}", bold=True)
        _fmt_cell(ws, f"B{r}")
        _fmt_cell(ws, f"C{r}", bold=True, align="right")
        _fmt_cell(ws, f"D{r}", bold=True, align="right", number_format="0.00%")
        r += 1
    return r


def _write_item_table(ws, r, name, rows):
    ws.cell(row=r, column=1, value=name)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _fmt_cell(ws, f"A{r}", bold=True, fill=SECTION_FILL)
    r += 1
    ws.cell(row=r, column=1, value="類別")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=3, value="人數")
    ws.cell(row=r, column=4, value="佔比")
    _fmt_cell(ws, f"A{r}", bold=True, fill=SUBHEADER_FILL, color="FFFFFF", align="center")
    _fmt_cell(ws, f"B{r}", fill=SUBHEADER_FILL)
    _fmt_cell(ws, f"C{r}", bold=True, fill=SUBHEADER_FILL, color="FFFFFF", align="center")
    _fmt_cell(ws, f"D{r}", bold=True, fill=SUBHEADER_FILL, color="FFFFFF", align="center")
    r += 1
    total = sum(cnt for _, cnt, _ in rows)
    for label, cnt, pct in rows:
        ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=3, value=cnt)
        ws.cell(row=r, column=4, value=pct)
        _fmt_cell(ws, f"A{r}")
        _fmt_cell(ws, f"B{r}")
        _fmt_cell(ws, f"C{r}", align="right")
        _fmt_cell(ws, f"D{r}", align="right", number_format="0.00%")
        r += 1
    # Total
    ws.cell(row=r, column=1, value="Total")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=3, value=total)
    ws.cell(row=r, column=4, value=1.0 if total else 0)
    for i in (1, 2, 3, 4):
        _fmt_cell(ws, f"{get_column_letter(i)}{r}", bold=True, fill=SUBTOTAL_FILL, align="right" if i >= 3 else "left")
    ws.cell(row=r, column=4).number_format = "0.00%"
    r += 1
    return r


def write_section2(ws, start_row: int, blocks_20, df_10) -> int:
    r = start_row
    ws.cell(row=r, column=1, value=f"區塊 2:ACC 1.0 vs ACC 2.0 對比 (1.0 n={len(df_10)}, 2.0 n=115)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    _fmt_cell(ws, f"A{r}", bold=True, fill=HEADER_FILL, color="FFFFFF", align="center")
    for col in range(1, 8):
        ws.cell(row=r, column=col).border = THICK_BORDER
    r += 2

    # 錄取國家對比
    r = _write_country_compare(ws, r, df_10, blocks_20)
    r += 1

    for item_20, col_10, custom in COMPARISON_MAP:
        rows_20 = blocks_20.get(item_20, [])
        if col_10 not in df_10.columns:
            print(f"  [警告] 1.0 找不到欄位: {col_10},跳過 {item_20}")
            continue
        if custom == "brand_age":
            rows_10 = count_10_brand_age(df_10, col_10)
        else:
            rows_10 = count_column(df_10, col_10)
        r = _write_compare_table(ws, r, item_20, rows_10, rows_20)
        r += 1

    # 備註
    r += 1
    ws.cell(row=r, column=1, value="備註:")
    _fmt_cell(ws, f"A{r}", bold=True)
    r += 1
    notes = [
        "- 1.0 站點欄若同一格多值 (例「日本,中東」),各國家 +1,分母用 1.0 賣家數,總和可能 >100%。",
        "- 1.0「US vs Expansion」依每位賣家站點是否含 US 判斷,屬互斥分類,加總=100%。",
        "- 1.0 無「公司所在區域」「Category」「預計開始銷售的時間」對應欄位,此次僅列 2.0 數據。",
        "- 註冊商標題目版本不同,本次不做對比。",
        "- 類別文字以各自原始值呈現,未做標準化映射,相似類別請逕行比較。",
    ]
    for n in notes:
        ws.cell(row=r, column=1, value=n)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        _fmt_cell(ws, f"A{r}")
        r += 1
    return r


def _write_country_compare(ws, r, df_10, blocks_20):
    ws.cell(row=r, column=1, value="錄取國家 (國家總計) + US vs Expansion")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    _fmt_cell(ws, f"A{r}", bold=True, fill=SECTION_FILL)
    r += 1
    # 欄位: 類別 | 1.0 人數 | 1.0 佔比 | | 2.0 人數 | 2.0 佔比 | 差異(百分點)
    headers = ["類別", "1.0 人數", "1.0 佔比", "", "2.0 人數", "2.0 佔比", "Δ 百分點"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
        _fmt_cell(ws, f"{get_column_letter(i)}{r}", bold=True, fill=SUBHEADER_FILL, color="FFFFFF", align="center")
    r += 1

    # 2.0 國家總計 (從 ae_rows 的小計抽出) — 直接用原本的錄取國家區塊重建
    ae_rows, us_exp_20 = get_recruit_country()
    country_20 = {}
    for country, bd, cnt, pct in ae_rows:
        if bd == "" and country.endswith("小計"):
            country_20[country.replace(" 小計", "")] = (cnt, pct)

    country_10 = {k: (c, p) for k, c, p in count_10_country(df_10)}

    for country in ["US", "MENA", "JP", "EU", "AU"]:
        c10, p10 = country_10.get(country, (0, 0))
        c20, p20 = country_20.get(country, (0, 0))
        diff = (p20 - p10) * 100
        _write_compare_row(ws, r, country, c10, p10, c20, p20, diff)
        r += 1
    # 額外列出 1.0 有但 2.0 沒有的區域 (如 SG)
    for extra in sorted(set(country_10) - {"US", "MENA", "JP", "EU", "AU"}):
        c10, p10 = country_10[extra]
        _write_compare_row(ws, r, extra, c10, p10, 0, 0, (0 - p10) * 100)
        r += 1

    # US vs Expansion
    ws.cell(row=r, column=1, value="US vs Expansion")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    _fmt_cell(ws, f"A{r}", bold=True, fill=EXPANSION_FILL)
    r += 1
    us_exp_10 = count_10_us_vs_expansion(df_10)
    map10 = {k: (c, p) for k, c, p in us_exp_10}
    map20 = {k: (c, p) for k, c, p in us_exp_20}
    for label in ["US", "Expansion (non-US)"]:
        c10, p10 = map10.get(label, (0, 0))
        c20, p20 = map20.get(label, (0, 0))
        diff = (p20 - p10) * 100
        _write_compare_row(ws, r, label, c10, p10, c20, p20, diff, bold=True)
        r += 1
    return r


def _write_compare_row(ws, r, label, c10, p10, c20, p20, diff, bold=False):
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=c10)
    ws.cell(row=r, column=3, value=p10)
    ws.cell(row=r, column=4, value="")
    ws.cell(row=r, column=5, value=c20)
    ws.cell(row=r, column=6, value=p20)
    ws.cell(row=r, column=7, value=diff)
    _fmt_cell(ws, f"A{r}", bold=bold)
    _fmt_cell(ws, f"B{r}", bold=bold, align="right")
    _fmt_cell(ws, f"C{r}", bold=bold, align="right", number_format="0.00%")
    _fmt_cell(ws, f"D{r}")
    _fmt_cell(ws, f"E{r}", bold=bold, align="right")
    _fmt_cell(ws, f"F{r}", bold=bold, align="right", number_format="0.00%")
    _fmt_cell(ws, f"G{r}", bold=bold, align="right", number_format='+0.0"pp";-0.0"pp";0"pp"')


def _write_compare_table(ws, r, title, rows_10, rows_20):
    ws.cell(row=r, column=1, value=title)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    _fmt_cell(ws, f"A{r}", bold=True, fill=SECTION_FILL)
    r += 1
    headers = ["類別", "1.0 人數", "1.0 佔比", "", "2.0 人數", "2.0 佔比", "Δ 百分點"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
        _fmt_cell(ws, f"{get_column_letter(i)}{r}", bold=True, fill=SUBHEADER_FILL, color="FFFFFF", align="center")
    r += 1
    map10 = {k: (c, p) for k, c, p in rows_10}
    map20 = {k: (c, p) for k, c, p in rows_20}
    all_labels = list(map20.keys())  # 以 2.0 為主序
    for k in map10:
        if k not in all_labels:
            all_labels.append(k)
    for label in all_labels:
        c10, p10 = map10.get(label, (0, 0))
        c20, p20 = map20.get(label, (0, 0))
        diff = (p20 - p10) * 100
        _write_compare_row(ws, r, label, c10, p10, c20, p20, diff)
        r += 1
    return r


# ============================================================
# main
# ============================================================

def main() -> None:
    print("[1/4] 讀取 2.0 分析...")
    blocks_20 = get_20_blocks()
    for k, rows in blocks_20.items():
        print(f"  {k}: {len(rows)} 類,總計 {sum(c for _, c, _ in rows)}")

    print("[2/4] 讀取 2.0 錄取國家...")
    ae_rows, us_exp = get_recruit_country()
    print(f"  AE 列數: {len(ae_rows)}  US vs Expansion: {us_exp}")

    print("[3/4] 讀取 1.0 最終錄取名單...")
    df_10 = load_10_roster()
    # 過濾空列 (公司名稱空的)
    df_10 = df_10[df_10["公司名稱"].notna()].copy()
    print(f"  1.0 賣家數: {len(df_10)}")

    print("[4/4] 寫 Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "ACC 2.0 分析"

    # 欄寬
    widths = [28, 18, 12, 12, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    r = write_section1(ws, r, blocks_20, ae_rows, us_exp)
    r += 2
    r = write_section2(ws, r, blocks_20, df_10)

    wb.save(OUT_FILE)
    print(f"完成: {OUT_FILE}")


if __name__ == "__main__":
    main()
